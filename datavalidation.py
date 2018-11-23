
from openpyxl import load_workbook
import pandas as pd
import datetime
import inspect as ins
from os import path
import sys , re


def df_comparison_old(testcase_id, tc_id_data, source_df, target_df, pathname):
    try:
        book = load_workbook(pathname)
        writer = pd.ExcelWriter(pathname)
        writer.book = book
        if testcase_id in book.sheetnames:
            name = book.get_sheet_by_name(testcase_id)
            book.remove_sheet(name)

        targetPrimaryKey = re.split(',',tc_id_data['targetPrimaryKey'])
        sourcePrimaryKey = re.split(',', tc_id_data['sourcePrimaryKey'])

        if tc_id_data.get('targetPrimaryKey') != '':
            indexed_df = target_df.set_index(targetPrimaryKey)
        else:
            print '*#*#*#*#*#*#TARGET PRIMARY KEY INVALID*#*#*#*#*#*#'

        if tc_id_data.get('sourcePrimaryKey') != '':
            indexed_df2 = source_df.set_index(sourcePrimaryKey)
        else:
            print '*#*#*#*#*#*#SOURCE PRIMARY KEY INVALID*#*#*#*#*#*#'


        for column in indexed_df:
            if column in tc_id_data['sourceColumn']:
                # moves the control back to the top of the loop
                continue
            else:
                indexed_df = source_df.drop(column, axis=1)
                # print 'dropped {}'.format(column)
        for column in indexed_df2:
            if column in tc_id_data['targetColumn']:
                # moves the control back to the top of the loop
                continue
            else:
                indexed_df2 = target_df.drop(column, axis=1)
                # print 'dropped {}'.format(column)

        diff_panel = pd.Panel(dict(df1=indexed_df2, df2=indexed_df))

        # Applying the diff function
        diff_output = diff_panel.apply(report_diff, axis=0)

        # Flag all the changes
        diff_output['has_change'] = diff_output.apply(has_change, axis=1)

        output_reduce = diff_output[(diff_output.has_change == 'Y')]


        if len(output_reduce) !=0:
            output_reduce.head(n=100).to_excel(writer, sheet_name=str(testcase_id), startrow=3,index= False)
            sheet = book.get_sheet_by_name(str(testcase_id))
            max_index = sheet.max_row
            sheet['A1'].value = 'Source ----> Target'
            sheet['C2'].value = "DATA COMPARISON"
            sheet['A{}'.format(max_index + 1)].value = 'Total Failure Count:'
            sheet['B{}'.format(max_index + 1)].value = len(output_reduce)
            max_index = sheet.max_row
            sheet['A{}'.format(max_index + 1)].value = 'Execution TimeStamp'
            sheet['B{}'.format(max_index + 1)].value = datetime.datetime.now()

        # red_text = Font(color="9C0006")
        # red_fill = PatternFill(bgColor="FFC7CE")
        # dxf = DifferentialStyle(font=red_text, fill=red_fill)
        # rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
        # rule.formula = ['NOT(ISERROR(SEARCH("--->")))']
        # writer.conditional_formatting.add('A1:F40', rule)

        # writer.save()
        if len(output_reduce) == 0:
            diff_output.head(n=100).to_excel(writer, sheet_name=str(testcase_id), startrow=3, index=False)
            sheet = book.get_sheet_by_name(str(testcase_id))
            max_index = sheet.max_row
            sheet['A{}'.format(max_index + 1)].value = 'Execution TimeStamp'
            sheet['B{}'.format(max_index + 1)].value = datetime.datetime.now()
            writer.save()
            print "No Mismatch Found, Verification Count as 0 == {}".format(len(output_reduce))
            return True
        else:
            print '#*#*#*#*#*#*#*#For {} , Number of Mismatch Found:{}'.format(testcase_id, len(output_reduce))
            writer.save()
            return False
    except Exception as e:
        print e
        print ins.stack()[0][3]
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        return False


def report_diff(x):
    # diff function to show the changes in each field
    return x[0] if x[0] == x[1] else '{} ---> {}'.format(*x)


def has_change(row):
    # tell which rows have changes
    if "--->" in row.to_string():
        return "Y"
    else:
        return "N"


def df_comparison(testcase_id, tc_id_data, source_df, target_df, pathname, tablesourcetarget):
    try:
        book = load_workbook(pathname)
        writer = pd.ExcelWriter(pathname)
        writer.book = book
        if testcase_id in book.sheetnames:
            del book[testcase_id]
            print "*#*#*#*#*#*#*# SHEET DELETED :{}*#*#*#*#*#*#*#".format(testcase_id)

        indexed_df = list(source_df)
        indexed_df2 = list(target_df)
        sourceLen = len(tc_id_data['sourceColumn'])
        targetLen = len(tc_id_data['targetColumn'])

        print 'Length : {},{}'.format(sourceLen,targetLen)

        if sourceLen > 0 and targetLen > 0 and sourceLen == targetLen :
            for column in indexed_df:
                if column in tc_id_data['sourceColumn']:
                    continue
                else:
                    source_df = source_df.drop(column, axis=1)
                    print 'Source Column dropped {}'.format(column)

            for column in indexed_df2:
                if column in tc_id_data['targetColumn']:
                    continue
                else:
                    target_df = target_df.drop(column, axis=1)
                    print 'Target Column dropped {}'.format(column)
        elif targetLen > 0:
            for column in indexed_df:
                if column in tc_id_data['targetColumn']:
                    continue
                else:
                    source_df = source_df.drop(column, axis=1)
                    print'Source Column dropped {}, picked from TargetColumn'.format(column)

            for column in indexed_df2:
                if column in tc_id_data['targetColumn']:
                    continue
                else:
                    target_df = target_df.drop(column, axis=1)
                    print'Target Column dropped {}'.format(column)

        else:
            print "*#*#*#*#*#*#*# NO TARGET COLUMNS ARGUMENT! CHECKING IN ALL DATA*#*#*#*#*#*#*#*#"

        diffOutput = pd.merge(source_df, target_df, how='outer', indicator=True)

        outputReduce = diffOutput[(diffOutput._merge != 'both')]
        outputReduce = outputReduce.replace({'_merge': {'left_only': 'SOURCE', 'right_only': 'TARGET'}})

        print outputReduce


        if len(outputReduce) !=0:
            outputReduce.head(n=100).to_excel(writer, sheet_name=str(testcase_id), startrow=6,index= False)
            #sheet = book.get_sheet_by_name(str(testcase_id))
            sheet = book[str(testcase_id)]
            sheet['A1'].value = '_merge '
            sheet['A2'].value = 'both'
            sheet['A3'].value = 'SOURCE'
            sheet['A4'].value = 'TARGET'
            sheet['B1'].value = 'SOURCE AND TARGET'
            sheet['B2'].value = 'DATA MATCHED'
            sheet['B3'].value = 'MISSING IN TARGET'
            sheet['B4'].value = 'MISSING IN SOURCE'
            sheet['C5'].value = "DATA COMPARISON"
            max_index = sheet.max_row
            sheet['A{}'.format(max_index + 1)].value = 'Total Failure Count:'
            sheet['B{}'.format(max_index + 1)].value = len(outputReduce)
            max_index = sheet.max_row
            sheet['A{}'.format(max_index + 1)].value = 'Execution TimeStamp'
            sheet['B{}'.format(max_index + 1)].value = datetime.datetime.now()

        if len(outputReduce) == 0:
            diffOutput.head(n=100).to_excel(writer, sheet_name=str(testcase_id), startrow=6, index=False)
            #sheet = book.get_sheet_by_name(str(testcase_id))
            sheet = book[str(testcase_id)]
            sheet['A1'].value = '_merge '
            sheet['A2'].value = 'both'
            sheet['A3'].value = 'SOURCE'
            sheet['A4'].value = 'TARGET'
            sheet['B1'].value = 'SOURCE AND TARGET'
            sheet['B2'].value = 'DATA MATCHED'
            sheet['B3'].value = 'MISSING IN TARGET'
            sheet['B4'].value = 'MISSING IN SOURCE'
            sheet['C5'].value = "DATA COMPARISON"
            max_index = sheet.max_row
            sheet['A{}'.format(max_index + 1)].value = 'Total Failure Count:'
            sheet['B{}'.format(max_index + 1)].value = len(outputReduce)
            max_index = sheet.max_row
            sheet['A{}'.format(max_index + 1)].value = 'Execution TimeStamp'
            sheet['B{}'.format(max_index + 1)].value = datetime.datetime.now()
            writer.save()
            print "No Mismatch Found, Verification Count as 0 == {}".format(len(outputReduce))
            return True
        else:
            print '#*#*#*#*#*#*#*#For {} , Number of Mismatch Found:{}'.format(testcase_id, len(outputReduce))
            writer.save()
            return False


    except Exception as e:
        print e
        print ins.stack()[0][3]
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        return False