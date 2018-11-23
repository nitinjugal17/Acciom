
import re, sys
from os import path
import inspect as ins
from openpyxl import load_workbook
import pandas as pd
import ux


def get_testdata(test_selected, df_testcase):
    df_data = pd.DataFrame()
    dict_data = {}
    tc_index = ''
    pd.set_option('max_colwidth', 1024)
    for tc_id in test_selected:
        df_data = df_testcase.loc[[tc_id]]
        checkColumn = ['Test Case ID','Columns','Details','Test Class','Test queries','TableSourceTarget']
        for column in checkColumn:
            if column in list(df_data):
                print "Column Exists:{}".format(column)
                continue
            else:
                print '*#*#*#*#*#Please Make Sure You Have Sheet with Column Name:{}*#*#*#*#*#'.format(column)

        str_tcid = df_data.to_string(columns=['Test Case ID'], index=False, header=False)
        str_data = df_data.to_string(columns=['Columns'], index=False, header=False)
        str_details = df_data.to_string(columns=['Details'], index=False, header=False)
        str_test_class = df_data.to_string(columns=['Test Class'], index=False, header=False)
        str_query = df_data.to_string(columns=['Test queries'], index=False, header=False)
        sourceTables = df_data.to_string(columns=['TableSource:Target'], index=False, header=False)
        #targetTables = df_data.to_string(columns=['TableTarget'], index=False, header=False)


        str_data = str_data.encode('utf8')
        str_tcid = str_tcid.encode('utf8')
        str_test_class = str_test_class.encode('utf8')
        str_query = str_query.encode('utf8')
        str_details = str_details.encode('utf8')
        sourceTables = sourceTables.encode('utf8')
        #targetTables = targetTables.encode('utf8')



        str_data = re.split('@|\n|\\n|[|]', str_data)
        str_details = re.split('@|\n|\\n|[|]', str_details)
        str_query = re.split('@|\\n|:', str_query)
        str_data.insert(0, str_tcid)
        str_data.insert(1, str_test_class)
        str_data.extend(str_details)

        # sourceTables = re.sub(r'[\n|\s+|\\\\]*','',sourceTables)
        # targetTables = re.sub('[\n]*','',targetTables)
        sourceTables = sourceTables.replace('\n','')
        sourceTables = re.split('[;]', sourceTables)
        #targetTables = re.split('[;]', targetTables)
        targetTables = []
        for i in range(0,len(sourceTables)) :
            strCopy = sourceTables[i]

            strCopy =  strCopy.strip('\\n')
            if not strCopy == '' :
                strCopy = strCopy.split(':')
                sourceTables[i] = strCopy[0]
                targetTables.insert(i,strCopy[1])

        sourceTables = [x for x in sourceTables if x]   # list comprehension for removing empty string

        #for i in range(0, len(targetTables)):
         #   strCopy = targetTables[i]
        #  strCopy = strCopy.strip('\\n')
         #   targetTables[i] = strCopy

        targetTables = [x for x in targetTables if x]  # list comprehension for removing empty string

        print sourceTables,targetTables



        for i in range(0,len(targetTables)):

            for itr in str_data:

                itr = itr.encode('unicode_escape')
                itr = re.sub('[\s+]', '', itr)        # for replacing whitespace space
                itr = re.sub(r'[\n]*', '', itr)       # for replacing newline comment space
                itr = re.split(':|\s|\\\\', itr)

                # Check Testcase names starts with TC_ (hard-coded)
                if "TC_" in itr[0]:
                    test_class = str_data[1]
                    dict_data[itr[0] + '_' + test_class + '-' + str(i)] = {}
                    tc_index = itr[0] + '_' + test_class + '-' + str(i)
                    #dict_data[itr[0] + '_' + test_class] = {}
                    #tc_index = itr[0] + '_' + test_class
                    dict_data[tc_index]['testClass'] = test_class
                    if 'querySource' in str_query:
                        squery_index = str_query.index('querySource')
                        sourceQuery = str_query[squery_index + 1]
                        sourceQuery = sourceQuery.replace('\'', '')
                        dict_data[tc_index]['querySource'] = sourceQuery.replace('\\n', '')
                    else:
                        print "************ NO SOURCE QUERY********************"
                    if 'queryTarget' in str_query:
                        tquery_index = str_query.index('queryTarget')
                        targetQuery = str_query[tquery_index + 1]
                        dict_data[tc_index]['queryTarget'] = targetQuery.replace('\\n', '')
                    else:
                        print "************ NO TARGET QUERY********************"


                elif "sourcedbType" in itr:
                    itr = itr[1].replace('\\n', "")
                    itr = itr.replace('\n', "")
                    itr = itr.replace(':', "")
                    itr = itr.replace('\\', "")
                    dict_data[tc_index]['sourcedbType'] = itr

                elif "sourceServer" in itr:
                    itr = itr[1].replace('\\n', "")
                    itr = itr.replace('\n', "")
                    itr = itr.replace(':', "")
                    itr = itr.replace('\\', "")
                    dict_data[tc_index]['sourceServer'] = itr

                elif "sourcedb" in itr:
                    itr = itr[1].replace('\\n', "")
                    itr = itr.replace('\n', "")
                    itr = itr.replace(':', "")
                    itr = itr.replace('\\', "")
                    dict_data[tc_index]['sourcedb'] = itr

                # elif "sourceTable" in itr:
                #     itr = itr[1].replace('\\n', "")
                #     itr = itr.replace('\n', "")
                #     itr = itr.replace(':', "")
                #     itr = itr.replace('\\', "")
                #     dict_data[tc_index]['sourceTable'] = itr

                # elif "sourceQuery" in itr :
                #     itr = itr[1].replace('\\n', "")
                #     itr = itr.replace('\n', "")
                #     itr = itr.replace(':', "")
                #     itr = itr.replace('\\', "")
                #     dict_data[tc_index]['sourceQuery'] = itr

                elif "sourceColumn" in itr:
                    itr = itr[1].replace('\\n', "")
                    itr = itr.replace('\n', "")
                    itr = itr.replace(':', "")
                    itr = itr.replace('\\', "")
                    dict_data[tc_index]['sourceColumn'] = itr

                elif "targetdbType" in itr:
                    itr = itr[1].replace('\\n', "")
                    itr = itr.replace('\n', "")
                    itr = itr.replace(':', "")
                    itr = itr.replace('\\', "")
                    dict_data[tc_index]['targetdbType'] = itr

                elif "targetServer" in itr:
                    itr = itr[1].replace('\\n', "")
                    itr = itr.replace('\n', "")
                    itr = itr.replace(':', "")
                    itr = itr.replace('\\', "")
                    dict_data[tc_index]['targetServer'] = itr

                elif "targetdb" in itr:
                    itr = itr[1].replace('\\n', "")
                    itr = itr.replace('\n', "")
                    itr = itr.replace(':', "")
                    itr = itr.replace('\\', "")
                    dict_data[tc_index]['targetdb'] = itr

                # elif "targetTable" in itr:
                #     itr = itr[1].replace('\\n', "")
                #     itr = itr.replace('\n', "")
                #     itr = itr.replace(':', "")
                #     itr = itr.replace('\\', "")
                #     dict_data[tc_index]['targetTable'] = itr


                # elif "targetQuery" in itr:
                #     itr = itr[1].replace('\\n', "")
                #     itr = itr.replace('\n', "")
                #     itr = itr.replace(':', "")
                #     itr = itr.replace('\\', "")
                #     dict_data[tc_index]['targetQuery'] = itr

                elif "targetColumn" in itr:
                    itr = itr[1].replace('\\n', "")
                    itr = itr.replace('\n', "")
                    itr = itr.replace(':', "")
                    itr = itr.replace('\\', "")
                    dict_data[tc_index]['targetColumn'] = itr

                elif "sourcePrimaryKey" in itr:
                    itr = itr[1].replace('\\n', "")
                    itr = itr.replace('\n', "")
                    itr = itr.replace(':', "")
                    itr = itr.replace('\\', "")
                    dict_data[tc_index]['sourcePrimaryKey'] = itr

                elif "targetPrimaryKey" in itr:
                    itr = itr[1].replace('\\n', "")
                    itr = itr.replace('\n', "")
                    itr = itr.replace(':', "")
                    itr = itr.replace('\\', "")
                    dict_data[tc_index]['targetPrimaryKey'] = itr

                elif "excludeColumns" in itr:
                    itr = itr[1].replace('\\n', "")
                    itr = itr.replace('\n', "")
                    itr = itr.replace(':', "")
                    itr = itr.replace('\\', "")
                    dict_data[tc_index]['excludeColumns'] = itr

                elif targetTables[i] != '':
                    dict_data[tc_index]['targetTable'] = targetTables[i]
                    if len(sourceTables) >= i:
                        dict_data[tc_index]['sourceTable'] = sourceTables[i]



        # if 'MultiTable' in df_data.columns :
        #     str_tables = df_data.to_string(columns=['MultiTable'], index=False, header=False)
        #     str_tables = str_tables.encode('utf8')
        #     str_tables = re.split('@|\n|\\n|[|]', str_tables)
        #     #str_data.append(str_tables)
        #     str_data.insert(2,str_tables)

        print str_data,dict_data

    return dict_data
    # pd.reset_option('max_colwidth')


def update_result(testcase_id, pathname, selected_sheet, override=''):
    try:
        splitName = testcase_id.split('-')
        print '*#*#*#*#*#*#*#*#*#*# FOR TEST :{}*#*#*#*#*#*#*#*#*#*'.format(str(testcase_id))
        if "CountCheck" in splitName[0]:

            testcase_id = testcase_id[::-1]
            testcase_id = testcase_id.replace(splitName[-1], '0', 1)
            testcase_id = testcase_id[::-1]
            print'*#*#*#*#*#*#*#*#*#*# UPDATING Results in SHEET Name FOR TEST :{}*#*#*#*#*#*#*#*#*#*'.format(str(testcase_id))

        check_index_len = 0
        check_df = pd.read_excel(pathname, sheet_name=str(testcase_id))
        check_index_len += len(check_df.index)
        work_book = load_workbook(pathname)
        #sheet = work_book.get_sheet_by_name(selected_sheet)
        sheet = work_book[selected_sheet]
        len_row = sheet.max_row
        #sheet2 = work_book.get_sheet_by_name(str(testcase_id))
        sheet2 = work_book[str(testcase_id)]
        max_index = sheet2.max_row
        print max_index
        if sheet2['D{}'.format(max_index + 4)].value == 'PASS' or override == 'PASS':
            override = 'PASS'
        else:
            override = 'FAIL'

        if override == 'PASS':
            for cellObj in sheet['A1':'B{}'.format(len_row)]:
                for cell in cellObj:
                    # print(cell.coordinate, cell.value)
                    if str(cell.value) in testcase_id:
                        cell_index = map(int, re.findall('\d+', cell.coordinate))
                        sheet['B{}'.format(cell_index[0])].value = 'PASS'
                        print 'TEST STATUS MUST BE UPDATED AS PASS for {}'.format(testcase_id)
                    break
            work_book.save(pathname)
            return True
        else:
            for cellObj in sheet['A1':'B{}'.format(len_row)]:
                for cell in cellObj:
                    # print(cell.coordinate, cell.value)
                    if str(cell.value) in testcase_id:
                        cell_index = map(int, re.findall('\d+', cell.coordinate))
                        sheet['B{}'.format(cell_index[0])].value = 'FAIL'
                        print 'TEST STATUS MUST BE UPDATED AS FAIL for {}'.format(testcase_id)
                    break
                work_book.save(pathname)
            return False
    except Exception as e:
        print e
        print ins.stack()[0][3]
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        return False


def create_results_sheet(sheet, pathname):
    try:
        work_book = load_workbook(pathname)

        # sheet = re.split('-', sheet)
        # print sheet[0]
        # print sheetnames,type(sheetnames)


        # for sheet in sheetnames:
        # if str(sheet[0]) in work_book.sheetnames and int(sheet[1]) == 0:
        #     work_book.get_sheet_by_name(str(sheet)).title = str(sheet) + '_old'
        #     #name = work_book.get_sheet_by_name(str(sheet))
        #     #name = work_book[str(sheet)]
        #     #work_book.remove_sheet(name)
        #
        # if str(sheet[0]+'_old') in work_book.sheetnames:
        #     del work_book[str(sheet[0]+sheet[1]+'_old')]
        #     print '*#*#*#*#*#*#*#*#*#*# OLD SHEET DELETED FOR TEST :{}*#*#*#*#*#*#*#*#*#*'.format(str(sheet[0]+'_old'))
        #
        # if int(sheet[1]) == 0:
        #     work_book.create_sheet(str(sheet[0]))
        #     print '*#*#*#*#*#*#*#*#*#*# FRESH SHEET CREATED FOR TEST :{}*#*#*#*#*#*#*#*#*#*'.format(str(sheet[0]))
        # else:
        #     if work_book[str(sheet[0])]:
        #         continue
        #     else:
        #         work_book.create_sheet(str(sheet[0]))
        splitName = sheet.split('-')
        print '*#*#*#*#*#*#*#*#*#*# FOR TEST :{}*#*#*#*#*#*#*#*#*#*'.format(str(sheet))
        if "CountCheck" in splitName[0] :
            sheet = sheet[::-1]
            sheet = sheet.replace(splitName[-1], '0', 1)
            sheet = sheet[::-1]
            print '*#*#*#*#*#*#*#*#*#*#  SHEET Name FOR TEST :{}*#*#*#*#*#*#*#*#*#*'.format(str(sheet))

        if str(sheet) in work_book.sheetnames:
            if "CountCheck" in splitName[0]:
                print '*#*#*#*#*#*#*#*#*#*# ESCAPE SHEET CREATION FOR TEST :{}*#*#*#*#*#*#*#*#*#*'.format(str(sheet))
            else:
                del work_book[str(sheet)]
                print '*#*#*#*#*#*#*#*#*#*# OLD SHEET DELETED FOR TEST :{}*#*#*#*#*#*#*#*#*#*'.format(str(sheet))
                work_book.create_sheet(str(sheet))
                print '*#*#*#*#*#*#*#*#*#*# FRESH SHEET CREATED FOR TEST :{}*#*#*#*#*#*#*#*#*#*'.format(str(sheet))
        else:
            work_book.create_sheet(str(sheet))
            print '*#*#*#*#*#*#*#*#*#*# FRESH SHEET CREATED FOR TEST :{}*#*#*#*#*#*#*#*#*#*'.format(str(sheet))

        work_book.save(pathname)
        return True
    except Exception as e:
        print e
        print ins.stack()[0][3]
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        print e


def read_testcase(pathname):
    try:
        pd.set_option('max_colwidth', 1024)
        if pathname != '':
            df_excel_sheet = pd.ExcelFile(pathname)
            sheet_list = df_excel_sheet.sheet_names
            # calling ux.select_sheet to select the sheet for reading test case
            selected_sheet = ux.select_sheet(sheet_list)
            # selected_sheet = selected_sheet[0].encode('utf8')
            df_excel = pd.read_excel(pathname, sheet_name=selected_sheet)
        else:
            print 'No Pathname Available , Retry Selecting the File'
        # print df_excel
        return df_excel, selected_sheet
    except IOError:
        wx.LogError("Cannot open file '%s'." % file)
        print e
        print ins.stack()[0][3]
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)