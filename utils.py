# -*- coding: utf8 -*-
# Firebird
# Microsoft SQL Server
# MySQL
# Oracle
# PostgreSQL
# SQLite
# Sybase
# ibm_db_sa - driver for IBM DB2 and Informix, developed jointly by IBM and SQLAlchemy developers.
# sqlalchemy-redshift - driver for Amazon Redshift, adapts the existing PostgreSQL/psycopg2 driver.
# sqlalchemy_exasol - driver for EXASolution.
# sqlalchemy-sqlany - driver for SAP Sybase SQL Anywhere, developed by SAP.
# sqlalchemy-monetdb - driver for MonetDB.
# snowflake-sqlalchemy - driver for Snowflake.
# sqlalchemy-tds - driver for MS-SQL, on top of pythone-tds.
# crate - driver for CrateDB.

import pandas as pd
from os import path
import sys
import re
from openpyxl import load_workbook
import datetime
from utilsexcel import get_testdata, read_testcase, update_result
from utilspandas import create_datasource
from ux import open_filedialog, select_testcases, select_columns
from datavalidation import df_comparison
import inspect as ins
from nullcheck import check_null
from duplicatecheck import check_duplicates
from countcheck import check_count
from ddlcheck import dict_compare
import utilsexcel as ue


class App:

    def __init__(self):
        # selection of test case file , should be in xlsx format ,return path of file selected
        self.pathname = open_filedialog()

        # reading the excel file for the given location, returns selected sheet name and data as a data frame
        self.df_testcase, self.selected_sheet = read_testcase(self.pathname)

        # reading server credentials from file and return it as a dict
        if self.read_creds():

            self.dict_credits = self.read_creds()

            # using df_test data asking user to select the test cases needed to be executed and return as a list
            test_selected = select_testcases(self.df_testcase, self.pathname)

            # reading test data for selected test case and return the same as dict
            self.test_data = get_testdata(test_selected, self.df_testcase)

            # passing test data to do the further execution according to the test data provided
            self.testclass_selection(self.test_data)
        else:
            print "*#*#*#*#*#*#*#*#*PLEASE VERIFY YOUR CREDENTIAL FILES LOCATION / EXISTS AND RETRY !!!*#*#*#*#*#*#*#"

    def testclass_selection(self, test_data):
        result_dict = {}
        for key, value in test_data.iteritems():
            try:
                testcase_id = key
                result_dict[testcase_id] = 'NA'
                ue.create_results_sheet(testcase_id, self.pathname)
                source_df, target_df, source_meta, target_meta, tc_id_data ,tablesourcetarget = create_datasource\
                (testcase_id, self.test_data, self.dict_credits)


                if not target_df.empty:

                    if tc_id_data['testClass'] != 'CountCheck' and tc_id_data['testClass'] != 'Count Check' \
                            and tc_id_data['testClass'] != 'DDLCheck' and tc_id_data['testClass'] != 'DDL Check':
                        if tc_id_data['sourceColumn'] != '':
                            source_column = re.split('\[|\]|\,|\\\'|\\"', tc_id_data['sourceColumn'])
                            tc_id_data['sourceColumn'] = [x for x in source_column if x]

                        if tc_id_data['targetColumn'] != '':
                            target_column = re.split('\[|\]|\,|\\\'|\\"', tc_id_data['targetColumn'])
                            tc_id_data['targetColumn'] = [x for x in target_column if x]
                        else:
                            tc_id_data['targetColumn'] = select_columns(tc_id_data['excludeColumns'],
                                                                        tc_id_data['testClass'], target_df)
                    else:
                        print "COLUMN INFORMATION NOT NEEDED FOR EXECUTION TO PROCEED for {}".format(testcase_id)

                    if tc_id_data['testClass'] == 'DuplicateCheck' or tc_id_data['testClass'] == 'Duplicate Check':

                        if tc_id_data['targetColumn']:
                            if check_duplicates(tc_id_data['targetColumn'], testcase_id, target_df, self.pathname, tablesourcetarget):
                                update_result(testcase_id, self.pathname, self.selected_sheet, override='PASS')
                                result_dict[testcase_id] = 'SHOULD BE MARKED AS PASS'
                                print "{} Executed and Results as PASS  ".format(testcase_id)
                            else:
                                result_dict[testcase_id] = 'SHOULD BE MARKED AS FAIL'
                                update_result(testcase_id, self.pathname, self.selected_sheet, override='FAIL')
                                print "#*#*#*#*#*#*#*#*#*#*#*Something went Wrong WHILE EXECUTION {}, Results as FAIL "\
                            " *##*#*#*#*#*#*#*#*#".format(testcase_id)
                        else:
                            result_dict[testcase_id] = 'FAIL AS NO COLUMN GIVEN OR SELECTED FOR TEST'
                            update_result(testcase_id, self.pathname, self.selected_sheet, override='FAIL')
                            print "#*#*#*#*#*#*#* No Columns Selected for {} , Please Retry ! #*#*#*#*#*#*#*# "\
                                    .format(testcase_id)

                    elif tc_id_data['testClass'] == 'NullCheck' or tc_id_data['testClass'] == 'Null Check':

                        if tc_id_data['targetColumn']:

                            if check_null(tc_id_data['targetColumn'], testcase_id, target_df, self.pathname, tablesourcetarget):
                                update_result(testcase_id, self.pathname, self.selected_sheet, override='PASS')
                                result_dict[testcase_id] = 'SHOULD BE MARKED AS PASS'
                                print "{} Executed and Results as PASS  ".format(testcase_id)

                            else:
                                result_dict[testcase_id] = 'SHOULD BE MARKED AS FAIL'
                                update_result(testcase_id, self.pathname, self.selected_sheet,  override='FAIL')
                                print "{} Executed and Results as FAIL  ".format(testcase_id)
                        else:
                            result_dict[testcase_id] = 'FAIL AS NO COLUMN GIVEN OR SELECTED FOR TEST'
                            update_result(testcase_id, self.pathname, self.selected_sheet, override='FAIL')
                            print "#*#*#*#*#*#*#* No Columns Selected for {} , Please Retry ! #*#*#*#*#*#*#*# " \
                                .format(testcase_id)

                    else:
                        print "Please Add a \"Test Class\" Field for framework to know how to Execute Case "
                        print 'Ignore Above Log , if Selected DataValidation/CountCheck/DDLCheck'

                    if not source_df.empty:
                        if tc_id_data['testClass'] == 'DataValidation' or tc_id_data['testClass'] == 'Data Validation':

                            if tc_id_data['targetColumn']:
                                if df_comparison(testcase_id, tc_id_data, source_df, target_df, self.pathname, tablesourcetarget):
                                    update_result(testcase_id, self.pathname, self.selected_sheet, override='PASS')
                                    result_dict[testcase_id] = 'SHOULD BE MARKED AS PASS'
                                    print "{} Executed and Results as PASS  ".format(testcase_id)
                                else:
                                    result_dict[testcase_id] = 'SHOULD BE MARKED AS FAIL'
                                    update_result(testcase_id, self.pathname, self.selected_sheet, override='FAIL')
                                    print "#*#*#*#*#*#*#*#*#*#*#*Something went Wrong WHILE EXECUTION {}, Results as FAIL " \
                                    " *##*#*#*#*#*#*#*#*#".format(testcase_id)
                            else:
                                result_dict[testcase_id] = 'FAIL AS NO COLUMN GIVEN OR SELECTED FOR TEST'
                                update_result(testcase_id, self.pathname, self.selected_sheet, override='FAIL')
                                print "#*#*#*#*#*#*#* No Columns Selected for {} , Please Retry ! #*#*#*#*#*#*#*# " \
                                    .format(testcase_id)
                            continue

                        elif tc_id_data['testClass'] == 'CountCheck' or tc_id_data['testClass'] == 'Count Check':
                            # return true when completed else return false

                            if check_count(testcase_id, source_df, target_df, self.pathname, tablesourcetarget):
                                if update_result(testcase_id, self.pathname, self.selected_sheet):
                                    result_dict[testcase_id] = 'SHOULD BE MARKED AS PASS'
                                    print "{} Executed and Results as PASS  ".format(testcase_id)
                                else:
                                    result_dict[testcase_id] = 'SHOULD BE MARKED AS FAIL'
                                    print "{} Executed and Results as FAIL  ".format(testcase_id)
                            else:
                                result_dict[testcase_id] = 'SHOULD BE MARKED AS FAIL'
                                update_result(testcase_id, self.pathname, self.selected_sheet, override='FAIL')
                                print "{} Executed and Results as FAIL  ".format(testcase_id)
                            continue

                        elif tc_id_data['testClass'] == 'DDLCheck' or tc_id_data['testClass'] == 'DDL Check':
                            if dict_compare(source_meta, target_meta, testcase_id, self.pathname):
                                update_result(testcase_id, self.pathname, self.selected_sheet, override='PASS')
                                result_dict[testcase_id] = 'SHOULD BE MARKED AS PASS'
                                print "{} Executed and Results as PASS  ".format(testcase_id)
                            else:
                                update_result(testcase_id, self.pathname, self.selected_sheet, override='FAIL')
                                result_dict[testcase_id] = 'SHOULD BE MARKED AS FAIL'
                                print "#*#*#*#*#*#*#*#*#*#*#*Something went Wrong WHILE EXECUTION {}, Results as FAIL " \
                                " *##*#*#*#*#*#*#*#*#".format(testcase_id)
                            continue

                        else:
                            print"Please Add a \"Test Class\" Field for framework to know how to Execute Case"
                            continue
                    else:
                        result_dict[testcase_id] = 'Source DB Information is empty ! Verification Count :{}'.format(len(source_df))
                        print 'Source DB Information is empty ! Verification Count :{}'.format(len(source_df))

                else:
                    result_dict[testcase_id] = ' Target DB Information is empty ! Verification Count :{}'.format(len(target_df)) + '\n' \
                                               + 'Source DB Information is empty ! Verification Count :{}'.format(len(source_df))
                    print 'Target DB Information is empty ! Verification Count :{}'.format(len(target_df))
                    print 'Source DB Information is empty ! Verification Count :{}'.format(len(source_df))


            except Exception as e:
                print e
                print ins.stack()[0][3]
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)
                pass

        print '\n*-*-*-*-*-*-*-*-*-*FINAL RESULTS OF EXECUTION STARTS*-*-*-*-*-*-*-*-*-*\n'
        for tc_id, stat in result_dict.iteritems():
            print "{}   :   {}".format(tc_id, stat)
        print '\n*-*-*-*-*-*-*-*-*-*FINAL RESULTS OF EXECUTION ENDS*-*-*-*-*-*-*-*-*-*'

    def read_creds(self):
        try:
            print(path.abspath('creds'))
            if path.exists("/Users/Jugal/Documents/GitHub/Acciom/creds") != '':
                dict_creds = {}
                with open("/Users/Jugal/Documents/GitHub/Acciom/creds") as f:
                    for line in f:

                        if line != '\n' and line != '':
                            (key, val) = line.split("=")

                            if 'serverName' in key:
                                dict_key = val.replace('\n', '')
                                dict_creds[dict_key] = {}
                                dict_creds[dict_key][key] = val.replace('\n', '')

                            elif line != '':
                                dict_creds[dict_key][key] = val.replace('\n', '')
                print dict_creds
                return dict_creds
            else:
                print 'File Not Available in Path , Retry Checking Creds File'


        except Exception as e:
            print e
            print ins.stack()[0][3]
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)



# -----------------------UNUSED DEF-------------------------------------

    def index_check(self):
        #unused method
        df = pd.concat([self.source_df, self.target_df])
        df = df.reset_index(drop=True)
        #group by
        df_gpby = df.groupby(list(df.index))
        # get index of unique records
        idx = [x[0] for x in df_gpby.groups.values() if len(x) == 1]
        #filter
        print 'Index Change Check Before Comparison:{}'.format(df.reindex(idx))
        return df.reindex(idx)

    def df_comparison_unused(self, testcase_id):
        try:

            book = load_workbook(self.pathname)
            writer = pd.ExcelWriter(self.pathname)
            writer.book = book
            if testcase_id in book.sheetnames:
                name = book.get_sheet_by_name(testcase_id)
                book.remove_sheet(name)
                # print 'Duplicate Sheet Deleted'

            df = self.source_df
            df2 = self.target_df
            primaryKey = self.targetPrimaryKey
            index_missing = df[~self.source_df[primaryKey].isin(self.target_df[primaryKey])]
            index_missing2 = df2[~self.target_df[primaryKey].isin(self.source_df[primaryKey])]

            if len(index_missing) == 0 and len(index_missing2) == 0 :
                # Create a panel of the two dataframes
                diff_panel = pd.Panel(dict(df1=self.source_df,df2=self.target_df))

                #Applying the diff function
                diff_output = diff_panel.apply(self.report_diff, axis=0)

                # Flag all the changes
                diff_output['has_change'] = diff_output.apply(self.has_change, axis=1)

                # writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                #Save the changes to excel but only include the columns changed

                diff_output[(diff_output.has_change == 'Y')].to_excel(writer,sheet_name=str(testcase_id),index=False,columns=list(diff_output))
                sheet = book.get_sheet_by_name(str(testcase_id))
                max_index = sheet.max_row
                # Make the text of the cell bold and italic
                cell = sheet['A{}'.format(max_index + 1)]
                cell.font = cell.font.copy(bold=True)
                sheet['A{}'.format(max_index + 1)].value = 'Execution TimeStamp'
                sheet['B{}'.format(max_index+1)].value = datetime.datetime.now()
                writer.save()

                print "ONE-TO-ONE COMPARISON DONE FOR : {}".format(testcase_id)

                return True
            else:
                print "ONE-TO-ONE COMPARISON DONE FOR : {}".format(testcase_id)
                # sheet = book.get_sheet_by_name(str(testcase_id))
                # max_index = sheet.max_row
                # cell = sheet['A{}'.format(max_index + 1)]
                # cell.font = cell.font.copy(bold=True)
                # sheet['A{}'.format(max_index + 1)].value = 'ROWS NOT EXISTS IN TARGET TABLE'
                # max_index = sheet.max_row
                index_missing.to_excel(writer, sheet_name=str(testcase_id), index=False,startrow=3)
                sheet = book.get_sheet_by_name(str(testcase_id))
                cell = sheet['A2']
                cell.font = cell.font.copy(bold=True)
                sheet['A2'].value = 'ROWS NOT EXISTS IN TARGET TABLE'
                max_index = sheet.max_row
                cell = sheet['A{}'.format(max_index + 1)]
                cell.font = cell.font.copy(bold=True)
                sheet['A{}'.format(max_index + 1)].value = 'ROWS NOT EXISTS IN SOURCE TABLE'
                max_index = sheet.max_row
                index_missing2.to_excel(writer, sheet_name=str(testcase_id), index=False,startrow=max_index+1)
                max_index = sheet.max_row
                sheet['A{}'.format(max_index + 1)].value = 'Execution TimeStamp'
                sheet['B{}'.format(max_index + 1)].value = datetime.datetime.now()
                writer.save()
                return False

        except Exception as e:
            print e
            return False
