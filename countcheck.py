
from openpyxl import load_workbook
import pandas as pd
import datetime
from os import path
import inspect as ins
import sys, re


def check_count(testcase_id, source_df, target_df, pathname ,tablesourcetarget):
    try:
        source_column = source_df.columns.values.tolist()
        source_row = source_df.index.values.tolist()
        target_column = target_df.columns.values.tolist()
        target_row = target_df.index.values.tolist()

        book = load_workbook(pathname)
        writer = pd.ExcelWriter(pathname)
        writer.book = book
        #sheet = book.get_sheet_by_name(str(testcase_id))
        splitText = testcase_id.split('-')
        testcase_id = testcase_id[::-1]
        testcase_id = testcase_id.replace(splitText[-1], '0', 1)
        testcase_id = testcase_id[::-1]
        print testcase_id
        sheet = book[testcase_id]
        max_index = sheet.max_row
        # Make the text of the cell bold and italic
        # cell = sheet['A{}'.format(max_index + 1)]
        #sheet['A{}'.format(max_index + 2)].value = 'Column'

        sheet['B{}'.format(max_index + 2)].value = 'Source'
        sheet['C{}'.format(max_index + 2)].value = 'Target'
        sheet['D{}'.format(max_index + 2)].value = 'Result'

        sheet['A{}'.format(max_index + 3)].value = 'Table Name'
        tablesourcetarget = tablesourcetarget.split(',')

        sheet['B{}'.format(max_index + 3)].value = tablesourcetarget[0]
        sheet['C{}'.format(max_index + 3)].value = tablesourcetarget[1]


        #sheet['B{}'.format(max_index + 2)].value = len(source_df.name)
        #sheet['C{}'.format(max_index + 2)].value = len(target_df.name)
        sheet['A{}'.format(max_index + 4)].value = 'Row'
        sheet['B{}'.format(max_index + 4)].value = len(source_row)
        sheet['C{}'.format(max_index + 4)].value = len(target_row)


        if source_row == target_row:
            sheet['D{}'.format(max_index + 4)].value = 'PASS'
        else:
            sheet['D{}'.format(max_index + 4)].value = 'FAIL'

        #sheet['B5'].value = 'SOURCE COLUMNS'
        # max_index = sheet.max_row
        # for column in range(len(source_column)):
        #     sheet['B{}'.format(max_index + 1)].value = source_column[column]
        #
        # #sheet['C5'].value = 'TARGET COLUMN'
        # for index in range(len(target_column)):
        #     sheet['C{}'.format(max_index + 1)].value = target_column[index]

        max_index = sheet.max_row
        sheet['A{}'.format(max_index + 1)].value = 'Execution TimeStamp'
        sheet['B{}'.format(max_index + 1)].value = datetime.datetime.now()
        writer.save()

        return True
    except Exception as e:
        print e
        print ins.stack()[0][3]
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        return False
