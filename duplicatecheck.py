
from openpyxl import load_workbook
import pandas as pd
import datetime
import inspect as ins
from os import path


def check_duplicates(column_name, testcase_id, target_df, pathname):
    try:
        column_name = [x.encode('utf-8') for x in column_name]
        print "Duplicate Check on Columns : {}".format(column_name)

        max_index = 1
        check_len = 0
        if len(target_df.index) != 0:
            book = load_workbook(pathname)
            sheet = book.get_sheet_by_name(str(testcase_id))
            sheet['A1'].value = 'Duplicate Check'

            if len(column_name) != 0:
                df_dupes = target_df[target_df.duplicated(subset=column_name, keep=False)]
                writer = pd.ExcelWriter(pathname)
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                for column in df_dupes:
                    if column in column_name:
                        continue
                    else:
                        df_dupes = df_dupes.drop(column, axis=1)
                df_dupes.to_excel(writer, sheet_name=str(testcase_id), startrow=max_index + 2)
                max_index = sheet.max_row
                check_len = + len(df_dupes)

            # Make the text of the cell bold and italic
            cell2 = sheet['A{}'.format(max_index + 1)]
            cell2.font = cell2.font.copy(bold=True)
            sheet['A{}'.format(max_index + 1)].value = 'Execution TimeStamp'
            sheet['B{}'.format(max_index + 1)].value = datetime.datetime.now()
            writer.save()
            if check_len == 0:
                print "No Mismatch Found, Verification Count as 0 == {}".format(check_len)
                return True
            else:
                print " Mismatch Found, Total Count  == {}".format(check_len)
                return False
        else:
            print 'FOR {} , NO DATA IN DATA FRAME, TOTAL INDEX COUNT : {}'.format(testcase_id, len(target_df.index))
            return False

    except Exception as e:
        print e
        print ins.stack()[0][3]
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        return False