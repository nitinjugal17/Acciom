import pandas as pd
from openpyxl import load_workbook
import datetime
from os import path
import inspect as ins
import sys


def dict_compare(source_meta, target_meta, testcase_id, pathname):
    try:
        ddl_dict = {}    # Source metadata
        ddl2_dict = {}   # Target metadata
        max_index = 1
        counter = 0
        book = load_workbook(pathname)
        writer = pd.ExcelWriter(pathname)
        sheet = book.get_sheet_by_name(str(testcase_id))
        writer.book = book
        sheet['A1'].value = 'DDL Check'
        sheet['A2'].value = 'Column Name'
        sheet['B2'].value = 'Data Type'
        sheet['C2'].value = 'Nullable'
        sheet['D2'].value = 'Column Missing '
        for i in source_meta:
            col = i['name']
            ddl_dict[col] = [i]

        for i2 in target_meta:
            col2 = i2['name']
            ddl2_dict[col2] = [i2]

        for key in ddl_dict.keys():

            if key in ddl2_dict.keys():
                if str(ddl_dict[key][0]['type']) == str(ddl2_dict[key][0]['type']):
                    if str(ddl_dict[key][0]['nullable']) == str(ddl2_dict[key][0]['nullable']):
                        # print "Name Type Nullable Comparison True : {}".format(ddl_dict[key])
                        continue
                    else:
                        max_index = sheet.max_row
                        cell2 = sheet['A{}'.format(max_index + 1)]
                        cell2.font = cell2.font.copy(bold=True)
                        sheet['A{}'.format(max_index + 1)].value = key
                        sheet['B{}'.format(max_index + 1)].value = str(ddl_dict[key][0]['type'])
                        sheet['C{}'.format(max_index + 1)].value = str(ddl_dict[key][0]['nullable'])
                        sheet['D{}'.format(max_index + 1)].value = 'IN Target'
                        writer.save()
                        counter += 1
                        print "Column Nullable Mismatch : {} \n with \n {}".format(ddl2_dict[key], ddl_dict[key])
                else:
                    max_index = sheet.max_row
                    cell2 = sheet['A{}'.format(max_index + 1)]
                    cell2.font = cell2.font.copy(bold=True)
                    sheet['A{}'.format(max_index + 1)].value = key
                    sheet['B{}'.format(max_index + 1)].value = str(ddl_dict[key][0]['type'])
                    sheet['C{}'.format(max_index + 1)].value = str(ddl_dict[key][0]['nullable'])
                    sheet['D{}'.format(max_index + 1)].value = 'IN Target'
                    writer.save()
                    counter += 1
                    print "Column Type Mismatch : {} \n with \n {}".format(ddl2_dict[key], ddl_dict[key])

            else:
                max_index = sheet.max_row
                cell2 = sheet['A{}'.format(max_index + 1)]
                cell2.font = cell2.font.copy(bold=True)
                sheet['A{}'.format(max_index + 1)].value = key
                sheet['B{}'.format(max_index + 1)].value = str(ddl_dict[key][0]['type'])
                sheet['C{}'.format(max_index + 1)].value = str(ddl_dict[key][0]['nullable'])
                sheet['D{}'.format(max_index + 1)].value = 'IN Target'
                writer.save()
                counter += 1
                print "Column Not Found in Target Database :{}".format(ddl_dict[key])

        for key in ddl2_dict.keys():

            if key in ddl_dict.keys():
                if str(ddl_dict[key][0]['type']) == str(ddl2_dict[key][0]['type']):
                    if str(ddl_dict[key][0]['nullable']) == str(ddl2_dict[key][0]['nullable']):
                        # print "Name Type Nullable Comparison True : {}".format(ddl_dict[key])
                        continue
                    else:
                        max_index = sheet.max_row
                        cell2 = sheet['A{}'.format(max_index + 1)]
                        cell2.font = cell2.font.copy(bold=True)
                        sheet['A{}'.format(max_index + 1)].value = key
                        sheet['B{}'.format(max_index + 1)].value = str(ddl2_dict[key][0]['type'])
                        sheet['C{}'.format(max_index + 1)].value = str(ddl2_dict[key][0]['nullable'])
                        sheet['D{}'.format(max_index + 1)].value = 'IN Source'
                        writer.save()
                        counter += 1
                        print "Column Nullable Mismatch : {} \n with \n {}".format(ddl2_dict[key], ddl_dict[key])
                else:
                    max_index = sheet.max_row
                    cell2 = sheet['A{}'.format(max_index + 1)]
                    cell2.font = cell2.font.copy(bold=True)
                    sheet['A{}'.format(max_index + 1)].value = key
                    sheet['B{}'.format(max_index + 1)].value = str(ddl2_dict[key][0]['type'])
                    sheet['C{}'.format(max_index + 1)].value = str(ddl2_dict[key][0]['nullable'])
                    sheet['D{}'.format(max_index + 1)].value = 'IN Source'
                    writer.save()
                    counter += 1
                    print "Column Type Mismatch : {} \n with \n {}".format(ddl2_dict[key], ddl_dict[key])

            else:
                max_index = sheet.max_row
                cell2 = sheet['A{}'.format(max_index + 1)]
                cell2.font = cell2.font.copy(bold=True)
                sheet['A{}'.format(max_index + 1)].value = key
                sheet['B{}'.format(max_index + 1)].value = str(ddl2_dict[key][0]['type'])
                sheet['C{}'.format(max_index + 1)].value = str(ddl2_dict[key][0]['nullable'])
                sheet['D{}'.format(max_index + 1)].value = 'IN Source'
                writer.save()
                counter += 1
                print "Column Not Found in Source Database :{}".format(ddl2_dict[key])

        max_index = sheet.max_row
        cell2 = sheet['A{}'.format(max_index + 1)]
        cell2.font = cell2.font.copy(bold=True)
        sheet['A{}'.format(max_index + 1)].value = 'Execution TimeStamp'
        sheet['B{}'.format(max_index + 1)].value = datetime.datetime.now()
        writer.save()

        if counter == 0:
            print "No Mismatch Found, Verification Count as 0 == {}".format(counter)
            return True
        else:
            print 'Total Number of Mismatch : {}'.format(counter)
            return False
    except Exception as e:
        print e
        print ins.stack()[0][3]
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
