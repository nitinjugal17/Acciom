
import pandas as pd
from openpyxl import load_workbook
import datetime
from os import path
import sys
import inspect as ins


def check_null(column_name, testcase_id, target_df, pathname):
    try:
        df = target_df
        nan_rows = pd.DataFrame()

        for column in column_name:
            nan_rows = nan_rows.append(df[target_df[column].isnull()])

        book = load_workbook(pathname)
        writer = pd.ExcelWriter(pathname)
        writer.book = book
        if testcase_id in book.sheetnames:
            #name = book.get_sheet_by_name(testcase_id)
            #book.remove_sheet(name)
            del book[str(testcase_id)]
        output_reduce = nan_rows.head(n=100)
        output_reduce.to_excel(writer, sheet_name=str(testcase_id), index=False, startrow=3)
        #sheet = book.get_sheet_by_name(str(testcase_id))
        sheet = book[str(testcase_id)]
        max_index = sheet.max_row
        sheet['A2'].value = "NULL CHECK --- Target DB"
        # Make the text of the cell bold and italic
        cell = sheet['A{}'.format(max_index + 1)]
        cell.font = cell.font.copy(bold=True)
        sheet['A{}'.format(max_index + 1)].value = 'Total Failure Count'
        sheet['B{}'.format(max_index + 1)].value = len(nan_rows)
        sheet['A{}'.format(max_index + 2)].value = 'Execution TimeStamp'
        sheet['B{}'.format(max_index + 2)].value = datetime.datetime.now()
        writer.save()
        if len(nan_rows) == 0:
            print "No Mismatch Found, Verification Count as 0 == {}".format(len(nan_rows))
            return True
        else:
            print '#*#*#*#*#*# FOR {} , NULL CHECK FAILED WITH TOTAL COUNT OF {}*#*#*#*#*#*#*#'\
                .format(testcase_id, len(nan_rows))
            return False
    except Exception as e:
        print e
        print ins.stack()[0][3]
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)